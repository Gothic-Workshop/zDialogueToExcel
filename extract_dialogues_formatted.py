
import os
import re
import yaml
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

CONFIG_FILE = "config.yaml"
INPUT_FOLDERS = ["Input_Gothic", "Input_Gothic2", "Input_NotR"]
OUTPUT_FOLDER = "Output"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

column_widths = {
    "A": 28.5,
    "B": 33.5,
    "C": 21.25,
    "D": 11.38,
    "E": 28.88
}

# Formatting styles
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
header_font = Font(bold=True)
bold_font = Font(bold=True)
italic_font = Font(italic=True)
self_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
other_fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
alignment = Alignment(wrap_text=True, vertical="center")

dotted = Side(style="dotted", color="000000")
thick = Side(style="medium", color="000000")

# Load config
if os.path.exists(CONFIG_FILE):
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        raw_mapping = yaml.safe_load(f)
        file_to_character = {k.lower(): v for k, v in raw_mapping.items()}
else:
    file_to_character = {}

# AI_Output pattern
ai_output_pattern = re.compile(
    r'AI_Output\s*\(\s*(\w+)\s*,\s*(\w+)\s*,\s*\"(.*?)\"\s*\)\s*;\s*(?:\/\/)?\s*(.*)'
)

columns = ["Output Name", "What character says", "Who says it", "Chapter", "Condition"]
character_data = defaultdict(lambda: defaultdict(list))

def get_dialogue_key(filename):
    return filename.replace(".d", "").replace("DIA_", "")

def normalize_character_name(dialogue_key):
    parts = dialogue_key.split("_")
    if len(parts) == 1 or not any(p.isdigit() for p in parts):
        return dialogue_key
    if parts[0].lower() == "addon" and len(parts) > 1:
        parts[1] = parts[1].upper()
        return "_".join(parts)
    parts[0] = parts[0].upper()
    return "_".join(parts)

# Gather files
file_instances = defaultdict(list)
for folder in INPUT_FOLDERS:
    for file in os.listdir(folder):
        if file.lower().endswith(".d"):
            file_instances[file.lower()].append((file, folder, os.path.join(folder, file)))

# Map filenames to characters
filename_to_character = {}
for file_lc, entries in file_instances.items():
    original_filename = entries[0][0]
    if file_lc in file_to_character:
        character_name = file_to_character[file_lc]
    else:
        key = get_dialogue_key(original_filename)
        character_name = normalize_character_name(key)
    filename_to_character[file_lc] = character_name

# Process files
for file_lc, entries in file_instances.items():
    character_name = filename_to_character[file_lc]
    for original_filename, folder, path in entries:
        sheet_name = folder.replace("Input_", "")
        with open(path, "r", encoding="windows-1252") as f:
            content = f.read()
        for speaker, listener, tag, line in ai_output_pattern.findall(content):
            character_data[character_name][sheet_name].append({
                "Output Name": tag,
                "What character says": line.strip(),
                "Who says it": speaker,
                "Chapter": "",
                "Condition": ""
            })

# Save files and format
for character, sheets in character_data.items():
    output_path = os.path.join(OUTPUT_FOLDER, f"{character}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet, rows in sheets.items():
            df = pd.DataFrame(rows, columns=columns)
            df.to_excel(writer, sheet_name=sheet[:31], index=False)

    wb = load_workbook(output_path)
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        max_row = sheet.max_row
        max_col = sheet.max_column

        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width

        # Header row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

        # Style data rows
        for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            who_cell = row[2]
            speaker = str(who_cell.value).strip().lower()
            row_fill = self_fill if speaker == "self" else other_fill if speaker in {"hero", "other"} else None

            for i, cell in enumerate(row):
                cell.alignment = alignment
                if i == 0:
                    cell.font = bold_font
                elif i == 1:
                    cell.font = italic_font
                if row_fill:
                    cell.fill = row_fill

        # Apply correct borders
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)

                border = Border(
                    top=thick if row_idx == 1 else dotted,
                    bottom=thick if row_idx == max_row else dotted,
                    left=thick if col_idx == 1 else dotted,
                    right=thick if col_idx == max_col else dotted,
                )
                cell.border = border

    wb.save(output_path)
    print(f"Saved: {output_path}")
